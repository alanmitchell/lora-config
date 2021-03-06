"""Utility routines related to sensors, needed by the configuration script.
"""

import json
import csv
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import config

def read_sensor_file(fn):
    """Reads a text file containing sensor key and model information, and returns
    a list of dictionaries containing the key information.  This routine attempts
    to automatically determine the type of input file, e.g. Elsys CSV file from the
    manufacturer.
    """
    with open(fn) as fin:
        first_line = fin.readline()
        if 'SKU;' in first_line and 'FW;' in first_line:
            return read_elsys_file(fn)

    raise ValueError('Unknown Sensor file type.')

def read_elsys_file(fn):
    """Reads an Elsys CSV file and returns a list of dictionaries containing key 
    info for each sensor.
    """
    recs = []

    with open(fn, newline='') as csvfile:
        reader = csv.DictReader(csvfile, delimiter=';')
        for row in reader:
            recs.append(dict(
                model = row['SKU'],
                dev_eui = row['EUI'],
                app_eui = row['AppEUI'],
                app_key = row["AppKey"]
            ))
        
    return recs

def make_things_v3_file(sensors):
    """Creates and returns the contents of a Things Stack version 3 JSON
    device import file, derived from the information in the list of 'sensors'.
    """
    contents = ''
    for s in sensors:
        dev_id = f"{s['model']}-{s['bldg_abbrev']}-{s['dev_eui'][-4:]}"
        dev_id = dev_id.replace('_', '-').replace(' ', '-').lower()
        rec = {
            'ids': {
                'device_id': dev_id,
                'dev_eui': s['dev_eui'],
                'join_eui': s['app_eui'],
            },
            'name': s['name'],
            'lorawan_version': s['lora_ver'][0],
            'lorawan_phy_version': s['lora_ver'][1],
            'frequency_plan_id': config.frequency_plan_id,
            'supports_join': True,
            'root_keys': {
                'app_key': {
                    'key': s['app_key'],
                },
            },
        }
        contents += json.dumps(rec) + '\n'
    
    return contents

def make_things_v2_file(sensors):
    """Creates and returns the contents of a file that can be used to Copy and Past values
    into the fields used for setting up a Things Network version 2 device.
    """
    contents = ''
    app_euis = set()

    for s in sensors:
        dev_id = f"{s['model']}-{s['bldg_abbrev']}-{s['dev_eui'][-4:]}"
        dev_id = dev_id.replace('_', '-').replace(' ', '-').lower()
        contents += f"Device ID:  {dev_id}\n"
        contents += f"Device EUI:  {s['dev_eui']}\n"
        contents += f"App Key:  {s['app_key']}\n"
        contents += f"Select this App EUI:  {s['app_eui']}\n"
        contents += f"{s['name']}\n\n"

        app_euis.add(s['app_eui'])

    contents = 'Prior to adding devices, make sure the following App EUIs are added to the Application:\n' + \
        '\n'.join(list(app_euis)) + '\n\n' + contents
    
    return contents

def write_bmon_spreadsheet(sensors, filename):
    """Saves an Excel spreadsheet to 'filename' that contains the neede sensor information
    to import into BMON.
    """

    columns = [
        ('Building', 23), 
        ('Sensor ID', 32), 
        ('Title', 28),
        ('Unit Label', 11), 
        ('Sensor Group', 29), 
        ('Sort Order', 8),
        ('Is Calc Field', 8),
        ('Calc or Transform Function', 24),
        ('Function Parameters', 33),
    ]

    # Make Workbook and get the worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Main'
    ws.append(list(zip(*columns))[0])

    # create all the sensor rows
    for s in sensors:
        for param, unit in s['params']:
            rec = [
                s['bldg'],
                f"{s['dev_eui']}_{param}",
                f"{s['name']} {config.param_label(param)}",
                unit,
            ]
            ws.append(rec)

    title = NamedStyle(name='title')
    title.font = Font(bold=True)
    bd = Side(style='thin', color='000000')
    title.border = Border(bottom=bd)
    title.alignment = Alignment(horizontal='center', wrap_text=True)
    wb.add_named_style(title)

    for col in 'ABCDEFGHI':
        ws[f'{col}1'].style = 'title'

    column_widths = list(zip(*columns))[1]
    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = column_width

    wb.save(filename)
